# -*- encoding: utf-8 -*-

"""
------------------------------------------
@File       : excel_utils.py
@Author     : maixiaochai
@Email      : maixiaochai@outlook.com
@CreatedOn  : 2022/2/28 14:15
------------------------------------------
"""
from openpyxl import load_workbook


class ExcelUtils:
    """
        处理 Excel的通用功能封装
    """

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = load_workbook(self.file_path)

    def read_sheet(self, sheet_index):
        """
        读取 sheet的文件
        :param sheet_index: sheet索引
        :return:            tuple
        """
        return self.workbook.worksheets[sheet_index].values

    def write_data(self, sheet_name, content: list):
        """
            写入数据
        :param sheet_name:      表单名称
        :param content:         内容
        """
        if self.exist_sheet(sheet_name):
            self.remove_sheet(sheet_name)
        sheet = self.workbook.create_sheet(sheet_name)

        for i in content:
            sheet.append(i)
        self.workbook.save(self.file_path)

    def exist_sheet(self, sheet_name) -> bool:
        """
        判断是否存在 名为 sheet_name的 sheet
        :param sheet_name:      sheet 的名称
        """
        return sheet_name in self.workbook.sheetnames

    def remove_sheet(self, sheet_name):
        """
        删除名为 sheet_name的 sheet
        """
        self.workbook.remove_sheet(self.workbook.get_sheet_by_name(sheet_name))
